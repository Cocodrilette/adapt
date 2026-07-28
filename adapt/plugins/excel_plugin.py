from __future__ import annotations
import logging
from adapt.cache import get_cache, set_cache, invalidate_cache

from itertools import zip_longest
from pathlib import Path
from typing import Any, Iterator, Sequence

from openpyxl import load_workbook

from .base import ResourceDescriptor, atomic_write
from .dataset_plugin import DEFAULT_HEADER_ROW, DatasetPlugin, resolve_header_row


logger = logging.getLogger(__name__)


def _is_blank(value: Any) -> bool:
    """True for a cell holding nothing a reader would call a value."""
    return value is None or (isinstance(value, str) and not value.strip())


def _merge_formula(value: Any, formula: Any) -> Any:
    """Fall back to a cell's formula text when it has no cached result.

    openpyxl reports only the value Excel last calculated and stored in the file.
    A workbook written by openpyxl itself carries no cached results at all, so
    under `data_only=True` every formula cell reads back as None and the column
    renders blank. Surfacing the formula is worse than a number and much better
    than an empty column with nothing to explain it.
    """
    if value is not None:
        return value
    if isinstance(formula, str) and formula.startswith("="):
        return formula
    return value


def _merged_rows(value_sheet: Any, formula_sheet: Any) -> Iterator[tuple[int, list[Any]]]:
    """Yield (1-based row number, cells) for a sheet, formulas filled in.

    Both sheets are streamed in lockstep so a large workbook is never fully
    materialised. `zip_longest` guards the case where the two passes disagree on
    width rather than silently truncating a row.
    """
    paired = zip_longest(
        value_sheet.iter_rows(values_only=True),
        formula_sheet.iter_rows(values_only=True),
        fillvalue=(),
    )
    for row_number, (value_row, formula_row) in enumerate(paired, start=1):
        yield row_number, [
            _merge_formula(value, formula)
            for value, formula in zip_longest(value_row or (), formula_row or ())
        ]


def _extract_header_and_sample(
    value_sheet: Any, formula_sheet: Any, header_row: int
) -> tuple[list[Any], list[Any]]:
    """Return the header cells and the first non-blank row beneath them."""
    header: list[Any] = []
    sample: list[Any] = []
    for row_number, cells in _merged_rows(value_sheet, formula_sheet):
        if row_number < header_row:
            continue
        if row_number == header_row:
            header = cells
            continue
        if all(_is_blank(cell) for cell in cells):
            continue
        sample = cells
        break
    return header, sample


class ExcelPlugin(DatasetPlugin):
    @property
    def resource_type(self) -> str:
        """Return the resource type string."""
        return "excel"

    def detect(self, path: Path) -> bool:
        """Detect if the path is an Excel file.

        Args:
            path: The file path to check.

        Returns:
            True if the file has .xlsx extension, False otherwise.
        """
        return path.suffix.lower() == ".xlsx"

    def load(self, path: Path) -> Sequence[ResourceDescriptor]:
        """Load Excel file and create descriptors for each worksheet.

        Every sheet is read at the default header row; a sheet that overrides it
        is re-read later by `apply_options`, once discovery has located the
        companion options file.

        Args:
            path: The path to the Excel file.

        Returns:
            A sequence of ResourceDescriptors, one for each worksheet.
        """
        logger.debug(f"Loading Excel file: {path}")
        descriptors = []
        with self._open_pair(path) as (values, formulas):
            for sheet_name in values.sheetnames:
                logger.debug(f"Processing sheet: {sheet_name}")
                descriptor = ResourceDescriptor(path=path, resource_type=self.resource_type)
                descriptor.metadata["primary_key"] = "_row_id"
                descriptor.metadata["sub_namespace"] = sheet_name
                self._set_header_metadata(
                    descriptor, values[sheet_name], formulas[sheet_name], DEFAULT_HEADER_ROW
                )
                descriptors.append(descriptor)
        logger.info(f"Loaded {len(descriptors)} worksheets from Excel file: {path}")
        return descriptors

    def apply_options(self, descriptor: ResourceDescriptor) -> None:
        """Honour a `header_row` override from the companion options file.

        Sheets often open with a title banner rather than column names, which
        would otherwise be parsed as the header and turn a sentence into a column
        name. Setting `{"header_row": 3}` points Adapt at the real header.
        """
        options = descriptor.metadata.get("options") or {}
        if "header_row" not in options:
            return

        descriptor.metadata["header_row"] = options["header_row"]
        header_row = resolve_header_row(descriptor)
        descriptor.metadata["header_row"] = header_row
        if header_row == DEFAULT_HEADER_ROW:
            return

        sheet_name = descriptor.metadata["sub_namespace"]
        logger.info(
            "Re-reading header of %s [%s] from row %d", descriptor.path, sheet_name, header_row
        )
        with self._open_pair(descriptor.path) as (values, formulas):
            self._set_header_metadata(
                descriptor, values[sheet_name], formulas[sheet_name], header_row
            )
        # Rows cached under the previous header row describe a different shape.
        invalidate_cache(str(descriptor.path))

    def _read_raw_rows(self, resource: ResourceDescriptor) -> list[list[str]]:
        """Read raw rows from the Excel worksheet.

        Rows that are entirely empty are dropped, and the sheet's extent is taken
        from the rows that actually carry values rather than from the declared
        `<dimension>`. Writers routinely leave styled-but-valueless cells behind
        (a whole formatted column, say), which inflates the dimension and would
        otherwise surface as hundreds of blank records. Note that openpyxl's
        `reset_dimensions()` is not a fix here — it falls back to the sheet
        maximum and yields over a million empty rows.

        Args:
            resource: The resource descriptor.

        Returns:
            A list of rows as lists of strings.
        """
        sub_namespace = resource.metadata["sub_namespace"]
        header_row = resolve_header_row(resource)
        logger.debug(f"Reading raw rows from Excel sheet: {sub_namespace}")
        cache_key = f"data:{resource.path}:{sub_namespace}:{header_row}"
        cached = get_cache(cache_key, str(resource.path))
        if cached:
            logger.debug(f"Cache hit for Excel data: {resource.path}")
            return cached

        logger.debug(f"Cache miss, reading from Excel file: {resource.path}")
        width = len(resource.metadata.get("header") or [])
        rows: list[list[str]] = []
        skipped = 0
        with self._open_pair(resource.path) as (values, formulas):
            merged = _merged_rows(values[sub_namespace], formulas[sub_namespace])
            for row_number, cells in merged:
                if row_number <= header_row:
                    continue
                if all(_is_blank(cell) for cell in cells):
                    skipped += 1
                    continue
                row = ["" if cell is None else str(cell) for cell in cells]
                if width:
                    row = (row + [""] * width)[:width]
                rows.append(row)

        if skipped:
            logger.debug(
                "Skipped %d empty rows in %s [%s]", skipped, resource.path, sub_namespace
            )
        set_cache(cache_key, rows, ttl_seconds=300, resource=str(resource.path))  # 5 min TTL
        return rows

    def _write_rows(self, resource: ResourceDescriptor, rows: list[dict[str, Any]], header: list[str]) -> None:
        """Write rows to the Excel worksheet.

        Args:
            resource: The resource descriptor.
            rows: The rows to write.
            header: The column headers.
        """
        sub_namespace = resource.metadata["sub_namespace"]
        header_row = resolve_header_row(resource)
        first_data_row = header_row + 1
        logger.info(f"Writing rows to Excel worksheet: {sub_namespace}")
        workbook = load_workbook(resource.path)
        try:
            sheet = workbook[sub_namespace]
            # Clear existing data, leaving the header and anything above it intact.
            for row in sheet.iter_rows(min_row=first_data_row):
                for cell in row:
                    cell.value = None
            # Write new data
            for row_idx, row_data in enumerate(rows, start=first_data_row):
                for col_idx, col_name in enumerate(header):
                    sheet.cell(row=row_idx, column=col_idx+1).value = row_data.get(col_name, "")

            def write_tmp(tmp_path: Path) -> None:
                workbook.save(tmp_path)

            atomic_write(resource.path, ".xlsx", write_tmp)
        finally:
            workbook.close()

        # Invalidate cache after mutation
        invalidate_cache(str(resource.path))
        logger.debug(f"Successfully wrote {len(rows)} rows to Excel file: {resource.path}")

    def _set_header_metadata(
        self, descriptor: ResourceDescriptor, value_sheet: Any, formula_sheet: Any, header_row: int
    ) -> None:
        """Parse the header and sample row for a sheet into descriptor metadata."""
        from .dataset_plugin import _ensure_header

        header, sample = _extract_header_and_sample(value_sheet, formula_sheet, header_row)
        descriptor.metadata["header"] = _ensure_header(header)
        descriptor.metadata["sample_row"] = [
            str(cell) if cell is not None else None for cell in sample
        ]
        descriptor.metadata["header_row"] = header_row

    @staticmethod
    def _open_pair(path: Path) -> "_WorkbookPair":
        """Open the workbook twice: once for cached values, once for formulas.

        Two passes are the only way to tell "empty cell" from "formula whose
        result was never cached" — openpyxl surfaces one or the other, never both.
        Both are read-only and streamed, and the result is cached for 5 minutes,
        so the second parse is paid rarely.
        """
        return _WorkbookPair(path)


class _WorkbookPair:
    """Context manager yielding (values_workbook, formulas_workbook)."""

    def __init__(self, path: Path) -> None:
        self._path = path

    def __enter__(self) -> tuple[Any, Any]:
        self._values = load_workbook(self._path, read_only=True, data_only=True)
        try:
            self._formulas = load_workbook(self._path, read_only=True, data_only=False)
        except Exception:
            self._values.close()
            raise
        return self._values, self._formulas

    def __exit__(self, *exc_info: Any) -> None:
        try:
            self._formulas.close()
        finally:
            self._values.close()
