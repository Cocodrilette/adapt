"""Tests for ExcelPlugin parsing: blank rows, sheet extent, formulas and header_row."""
import json
from pathlib import Path
import shutil
from types import SimpleNamespace

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from adapt.config import AdaptConfig
from adapt.discovery import discover_resources
from adapt.plugins.excel_plugin import ExcelPlugin


LEGACY_FIXTURE = Path(__file__).parent / "fixtures" / "legacy_inventory.xls"


@pytest.fixture(autouse=True)
def isolated_cache(tmp_path, monkeypatch):
    """Point the plugin cache at a per-test database."""
    from adapt import cache

    monkeypatch.setattr(cache, "_db_path", str(tmp_path / "cache.db"))


def _descriptor(plugin, path, sheet_name):
    for descriptor in plugin.load(path):
        if descriptor.metadata["sub_namespace"] == sheet_name:
            return descriptor
    raise AssertionError(f"sheet {sheet_name} not found")


def test_legacy_xls_is_discovered_and_read_per_sheet():
    """A BIFF8 workbook is exposed with typed rows instead of being rejected."""
    plugin = ExcelPlugin()

    assert plugin.detect(LEGACY_FIXTURE)
    descriptors = plugin.load(LEGACY_FIXTURE)
    assert [item.metadata["sub_namespace"] for item in descriptors] == [
        "Inventory",
        "Summary",
    ]

    inventory = _descriptor(plugin, LEGACY_FIXTURE, "Inventory")
    assert inventory.metadata["header"] == [
        "name",
        "quantity",
        "active",
        "updated_at",
    ]
    assert inventory.metadata["readonly"] is True

    request = SimpleNamespace(state=SimpleNamespace(user=None))
    assert plugin.read(inventory, request) == [
        {
            "_row_id": 1,
            "name": "Widget",
            "quantity": 12,
            "active": True,
            "updated_at": "2026-07-15 09:30:00",
        },
        {
            "_row_id": 2,
            "name": "Gadget",
            "quantity": 3,
            "active": False,
            "updated_at": "2026-07-16 14:00:00",
        },
    ]


def test_discovery_registers_legacy_xls_sheets(tmp_path):
    """The default registry loads .xls sheets during normal discovery."""
    root = tmp_path / "docroot"
    root.mkdir()
    shutil.copyfile(LEGACY_FIXTURE, root / "inventory.xls")

    resources = discover_resources(root, AdaptConfig(root=root))
    legacy_resources = [item for item in resources if item.path.suffix == ".xls"]

    assert [item.metadata["sub_namespace"] for item in legacy_resources] == [
        "Inventory",
        "Summary",
    ]
    assert all(item.metadata["readonly"] for item in legacy_resources)


def test_legacy_xls_mutation_is_rejected_before_file_access():
    """Adapt must not corrupt BIFF8 workbooks through an incompatible writer."""
    plugin = ExcelPlugin()
    inventory = _descriptor(plugin, LEGACY_FIXTURE, "Inventory")
    request = SimpleNamespace(state=SimpleNamespace(user=None))
    context = SimpleNamespace(readonly=False)

    with pytest.raises(HTTPException) as error:
        plugin.write(
            inventory,
            {"action": "create", "data": [{"name": "New"}]},
            request,
            context,
        )

    assert error.value.status_code == 405
    assert "convert the file to .xlsx" in error.value.detail.lower()


def test_empty_rows_are_dropped_wherever_they_appear(tmp_path):
    """A2: an all-empty row is not a record, leading, trailing or interior."""
    path = tmp_path / "gaps.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["name", "qty"])
    ws.append(["alpha", 1])
    ws.append([None, None])          # interior separator
    ws.append(["beta", 2])
    ws.append([None, None])          # trailing
    ws.append([None, None])
    wb.save(path)

    plugin = ExcelPlugin()
    rows = plugin._read_raw_rows(_descriptor(plugin, path, "Data"))

    assert rows == [["alpha", "1"], ["beta", "2"]]


def test_extent_comes_from_values_not_the_declared_dimension(tmp_path):
    """B: styled-but-valueless cells inflate <dimension> and must not become rows."""
    path = tmp_path / "styled.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["name", "qty"])
    ws.append(["alpha", 1])
    # Formatting applied far below the data, as a whole-column format would leave.
    for row in range(3, 200):
        ws.cell(row=row, column=2).font = Font(bold=True)
    wb.save(path)

    # The workbook really does claim to be 199 rows tall.
    probe = load_workbook(path, read_only=True)
    assert probe[ws.title].max_row >= 199
    probe.close()

    plugin = ExcelPlugin()
    rows = plugin._read_raw_rows(_descriptor(plugin, path, "Data"))

    assert rows == [["alpha", "1"]]


def test_rows_are_padded_to_the_header_width(tmp_path):
    """Short rows still carry every column, so the UI cannot see a missing key."""
    path = tmp_path / "ragged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["a", "b", "c"])
    ws.append(["only-a"])
    wb.save(path)

    plugin = ExcelPlugin()
    rows = plugin._read_raw_rows(_descriptor(plugin, path, "Data"))

    assert rows == [["only-a", "", ""]]


def test_formula_without_cached_value_falls_back_to_formula_text(tmp_path):
    """openpyxl-written workbooks cache no results; show the formula, not a blank."""
    path = tmp_path / "formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["label", "count"])
    ws.append(["total", "=COUNTA(Data!$A$2:$A$99)"])
    wb.save(path)

    plugin = ExcelPlugin()
    rows = plugin._read_raw_rows(_descriptor(plugin, path, "Data"))

    assert rows == [["total", "=COUNTA(Data!$A$2:$A$99)"]]


def test_row_of_only_uncached_formulas_survives_the_blank_filter(tmp_path):
    """Formulas are merged in before the empty-row test, or such a row would vanish."""
    path = tmp_path / "all_formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["x", "y"])
    ws.append(["=1+1", "=2+2"])
    wb.save(path)

    plugin = ExcelPlugin()
    rows = plugin._read_raw_rows(_descriptor(plugin, path, "Data"))

    assert rows == [["=1+1", "=2+2"]]


def test_cached_formula_value_is_preferred_over_the_formula(tmp_path):
    """When Excel has stored a result, that result is what the API serves."""
    path = tmp_path / "cached.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["label", "count"])
    ws.append(["total", "=1+1"])
    wb.save(path)

    # Simulate a calculating writer having stored the result.
    cached = load_workbook(path)
    cached["Data"]["B2"].value = 2
    cached.save(path)

    plugin = ExcelPlugin()
    rows = plugin._read_raw_rows(_descriptor(plugin, path, "Data"))

    assert rows == [["total", "2"]]


def _docroot_with(path_name: str, tmp_path: Path) -> tuple[Path, Path]:
    root = tmp_path / "docroot"
    root.mkdir()
    return root, root / path_name


def test_header_row_option_moves_the_header_off_a_title_banner(tmp_path):
    """C: `header_row` in the companion options file re-points the parser."""
    root, path = _docroot_with("report.xlsx", tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Pipeline Dashboard (auto-calculated)"])   # title banner
    ws.append([None, None])
    ws.append(["Stage", "Count"])                          # real header
    ws.append(["Active", 3])
    wb.save(path)

    options_dir = root / ".adapt"
    options_dir.mkdir()
    (options_dir / "report.Dashboard.options.json").write_text(
        json.dumps({"header_row": 3}), encoding="utf-8"
    )

    config = AdaptConfig(root=root)
    resources = discover_resources(root, config)
    resource = next(r for r in resources if r.metadata.get("sub_namespace") == "Dashboard")

    assert resource.metadata["header"] == ["Stage", "Count"]
    assert resource.metadata["header_row"] == 3

    plugin = ExcelPlugin()
    from adapt.plugins.base import ResourceDescriptor

    descriptor = ResourceDescriptor(
        path=resource.path,
        resource_type=resource.resource_type,
        schema_path=resource.schema_path,
        metadata=resource.metadata,
    )
    assert plugin._read_raw_rows(descriptor) == [["Active", "3"]]

    # The generated schema must describe the overridden header, not the banner.
    schema = json.loads(resource.schema_path.read_text(encoding="utf-8"))
    assert list(schema["columns"]) == ["Stage", "Count"]


def test_invalid_header_row_falls_back_to_the_default(tmp_path):
    """A typo in a hand-written options file must not break the resource."""
    root, path = _docroot_with("report.xlsx", tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name"])
    ws.append(["alpha"])
    wb.save(path)

    options_dir = root / ".adapt"
    options_dir.mkdir()
    (options_dir / "report.Sheet1.options.json").write_text(
        json.dumps({"header_row": "third"}), encoding="utf-8"
    )

    config = AdaptConfig(root=root)
    resources = discover_resources(root, config)
    resource = next(r for r in resources if r.metadata.get("sub_namespace") == "Sheet1")

    assert resource.metadata["header"] == ["name"]


def test_malformed_options_file_is_ignored(tmp_path):
    """A broken options file is logged and skipped, not fatal."""
    root, path = _docroot_with("report.xlsx", tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name"])
    ws.append(["alpha"])
    wb.save(path)

    options_dir = root / ".adapt"
    options_dir.mkdir()
    (options_dir / "report.Sheet1.options.json").write_text("{not json", encoding="utf-8")

    config = AdaptConfig(root=root)
    resources = discover_resources(root, config)
    resource = next(r for r in resources if r.metadata.get("sub_namespace") == "Sheet1")

    assert resource.metadata["header"] == ["name"]


def test_write_respects_an_overridden_header_row(tmp_path):
    """Writing must not clobber the title banner above an overridden header."""
    path = tmp_path / "report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Pipeline Dashboard (auto-calculated)"])
    ws.append([None, None])
    ws.append(["Stage", "Count"])
    ws.append(["Active", 3])
    wb.save(path)

    from adapt.plugins.base import ResourceDescriptor

    descriptor = ResourceDescriptor(
        path=path,
        resource_type="excel",
        metadata={
            "sub_namespace": "Dashboard",
            "header": ["Stage", "Count"],
            "header_row": 3,
            "primary_key": "_row_id",
        },
    )
    ExcelPlugin()._write_rows(
        descriptor, [{"Stage": "Awaiting", "Count": 7}], ["Stage", "Count"]
    )

    written = load_workbook(path)["Dashboard"]
    assert written["A1"].value == "Pipeline Dashboard (auto-calculated)"
    assert written["A3"].value == "Stage"
    assert written["A4"].value == "Awaiting"
    assert written["B4"].value == 7
